# -*- coding: utf-8 -*-
# /usr/bin/python
#########################################################
# Function :cangkujiance                                #
# Platform :UnionTech OS                                #
# Version  :1.0                                         #
# Date     :2022-08-29                                  #
# Author   :hanshuang                                   #
# Contact  :hanshuang@uniontech.com                     #
# Company  :UnionTech                                   #
#########################################################

# 导入openpyxl库
import os
import openpyxl
from openpyxl.styles import Alignment
import linecache
import re

path = './report/summary.out'
print(os.getcwd())
os.system('sh lmbench.sh')
def get_file_lmbench_value(keyword, downrow):
    with open(path, "r+") as f:
        readlines = f.readlines()
        number = 1
        for readline in readlines:
            if readline.find(keyword) != -1:
                break
            else:
                number += 1

    print("number = %d" %(number))
    row = number + downrow
    read_specified_line_value = linecache.getline(path, row)
    specified_line_value = read_specified_line_value.split()
    print("specified_line_value = %s\n" %(specified_line_value))
    return specified_line_value

def write_lmbench_excel(dict, row_1):
    ws.cell(row=row_1, column=1, value=dict['name'])
    ws.cell(row=row_1, column=2, value=dict['explain'])
    ws.cell(row=row_1, column=3, value=dict['value'])


# 新建一个工作薄
bw = openpyxl.Workbook()
# 激活表格
ws = bw.active
# 新建一个名称为lmbench_sheet1的表单
ws.title = 'lmbench_sheet1'
title = "Processor, Processes - times in microseconds - smaller is better       处理器Process性能"
# 在表格的第一行第一列插入变量名为title值
ws.cell(row=1, column=1, value=title)
title1 = "单位：μs"
# 在表格的第二行第一列插入变量名为title值
ws.cell(row=2, column=1, value=title1)
Processor_null_call = {'name':'获取进程ID的时延 null call', 
        'explain':'执行getppid需要的时间', 
        'keyword':'Processor, Processes',
        'downrow':5,
        }   
Processor_null_call['value'] = get_file_lmbench_value(Processor_null_call['keyword'], Processor_null_call['downrow'])[4]
# print("Processor_null_call['value'] = %s\n" %(Processor_null_call['value']))
Processor_null_IO = {'name':'系统调用读和写操作的平均时延 null I/O', 
        'explain':'从/dev/zero读一个字节的时间长', 
        'keyword':'Processor, Processes',
        'downrow':5,
        }   
Processor_null_IO['value'] = get_file_lmbench_value(Processor_null_IO['keyword'], Processor_null_IO['downrow'])[5]
# print("Processor_null_IO['value'] = %s\n" %(Processor_null_IO['value']))
Processor_stat = {'name':'系统调用获取文件信息时延 stat',   
        'explain':'得到一个文件的信息需用的时间',
        'keyword':'Processor, Processes',
        'downrow':5,
        }   
Processor_stat['value'] = get_file_lmbench_value(Processor_stat['keyword'], Processor_stat['downrow'])[6]
# print("Processor_stat['value'] = %s\n" %(Processor_stat['value']))
Processor_open_close = {'name':'系统调用打开/关闭时延 open close',
        'explain':'open一个文件然后再close它总共需用的时间（不包括读目录和节点的时间',
        'keyword':'Processor, Processes',
        'downrow':5,
        }
Processor_open_close['value'] = get_file_lmbench_value(Processor_open_close['keyword'], Processor_open_close['downrow'])[7]
# print("Processor_open_close['value'] = %s\n" %(Processor_open_close['value']))
Processor_slct_TCP = {'name':'系统调用监听文件描述符时延 slct TCP',
        'explain':'通过TCP网络连接选择100个文件描述符所耗用的时间',
        'keyword':'Processor, Processes',
        'downrow':5,
        }
Processor_slct_TCP['value'] = get_file_lmbench_value(Processor_slct_TCP['keyword'], Processor_slct_TCP['downrow'])[8]
# print("Processor_slct_TCP['value'] = %s\n" %(Processor_slct_TCP['value']))
Processor_sig_inst = {'name':'系统调用注册信号时延 sig inst',
        'explain':'install signal handler所耗用的时间',
        'keyword':'Processor, Processes',
        'downrow':5,
        }
Processor_sig_inst['value'] = get_file_lmbench_value(Processor_sig_inst['keyword'], Processor_sig_inst['downrow'])[9]
# print("Processor_sig_inst['value'] = %s\n" %(Processor_sig_inst['value']))
Processor_sig_hndl = {'name':'系统调用捕获信号时延 sig hndl',
        'explain':'catch signal 所耗用的时间',
        'keyword':'Processor, Processes',
        'downrow':5,
        }
Processor_sig_hndl['value'] = get_file_lmbench_value(Processor_sig_hndl['keyword'], Processor_sig_hndl['downrow'])[10]
# print("Processor_sig_hndl['value'] = %s\n" %(Processor_sig_hndl['value']))
Processor_fork_proc = {'name':'系统调用创建进程时延 fork proc',
        'explain':'fork一个完全相同的process，并把原来的process关掉所耗用的时间',
        'keyword':'Processor, Processes',
        'downrow':5,
        }
Processor_fork_proc['value'] = get_file_lmbench_value(Processor_fork_proc['keyword'], Processor_fork_proc['downrow'])[11]
# print("Processor_fork_proc['value'] = %s\n" %(Processor_fork_proc['value']))
Processor_exec_proc = {'name':'系统调用执行命令时延 exec proc',
        'explain':'fork一个新进程执行新命令，所耗用时间',
        'keyword':'Processor, Processes',
        'downrow':5,
        }
Processor_exec_proc['value'] = get_file_lmbench_value(Processor_exec_proc['keyword'], Processor_exec_proc['downrow'])[12]
# print("Processor_exec_proc['value'] = %s\n" %(Processor_exec_proc['value']))
Processor_sh_proc = {'name':'系统shell执行命令时延 sh proc',
        'explain':'fork一个新进程，同时询问系统shell来找到并运行一个新程序所耗用时间',
        'keyword':'Processor, Processes',
        'downrow':5,
        }
Processor_sh_proc['value'] = get_file_lmbench_value(Processor_sh_proc['keyword'], Processor_sh_proc['downrow'])[13]
# print("Processor_sh_proc['value'] = %s\n" %(Processor_sh_proc['value']))
row = 3
write_lmbench_excel(Processor_null_call, row)
row += 1
write_lmbench_excel(Processor_null_IO, row)
row += 1
write_lmbench_excel(Processor_stat, row)
row += 1
write_lmbench_excel(Processor_open_close, row)
row += 1
write_lmbench_excel(Processor_slct_TCP, row)
row += 1
write_lmbench_excel(Processor_sig_inst, row)
row += 1
write_lmbench_excel(Processor_sig_hndl, row)
row += 1
write_lmbench_excel(Processor_fork_proc, row)
row += 1
write_lmbench_excel(Processor_exec_proc, row)
row += 1
write_lmbench_excel(Processor_sh_proc, row)
row += 1

title11 = "Context switching - times in microseconds - smaller is better      上下文切换所花时间"
# 在表格的第一行第一列插入变量名为title值
ws.cell(row=row, column=1, value=title11)
row += 1
title12 = "单位：μs"
# 在表格的第二行第一列插入变量名为title值
ws.cell(row=row, column=1, value=title12)
Context_switching_2p_0k = {'name':'系统进程切换时延 2p/0K ctxsw',
        'explain':'每个进程size为0（不执行任何任务）进程数为2上下文切换耗时',
        'keyword':'Context switching',
        'downrow':5,
        }
Context_switching_2p_0k['value'] = get_file_lmbench_value(Context_switching_2p_0k['keyword'], Context_switching_2p_0k['downrow'])[3]
print("Context_switching_2p_0k['value'] = %s\n" %(Context_switching_2p_0k['value']))
Context_switching_2p_16K = {'name':'系统进程切换时延 2p/16K ctxsw',
        'explain':'每个进程size为16K（执行任务）进程数为2上下文切换耗时',
        'keyword':'Context switching',
        'downrow':5,
        }
Context_switching_2p_16K['value'] = get_file_lmbench_value(Context_switching_2p_16K['keyword'], Context_switching_2p_16K['downrow'])[4]
print("Context_switching_2p_16K['value'] = %s\n" %(Context_switching_2p_16K['value']))

Context_switching_2p_64K = {'name':'系统进程切换时延 2p/64K ctxsw',
        'explain':'每个进程size为64K（执行任务）进程数为2上下文切换耗时',
        'keyword':'Context switching',
        'downrow':5,
        }
Context_switching_2p_64K['value'] = get_file_lmbench_value(Context_switching_2p_64K['keyword'], Context_switching_2p_64K['downrow'])[5]
print("Context_switching_2p_64K['value'] = %s\n" %(Context_switching_2p_64K['value']))



Context_switching_8p_16K = {'name':'系统进程切换时延 8p/16K ctxsw',
        'explain':'每个进程size为16K（执行任务）进程数为8上下文切换耗时',
        'keyword':'Context switching',
        'downrow':5,
        }
Context_switching_8p_16K['value'] = get_file_lmbench_value(Context_switching_8p_16K['keyword'], Context_switching_8p_16K['downrow'])[6]
print("Context_switching__8p_16K['value'] = %s\n" %(Context_switching_8p_16K['value']))

Context_switching_8p_64K = {'name':'系统进程切换时延 8p/64K  ctxsw',
        'explain':'每个进程size为64K（执行任务）进程数为8上下文切换耗时',
        'keyword':'Context switching',
        'downrow':5,
        }
Context_switching_8p_64K['value'] = get_file_lmbench_value(Context_switching_8p_64K['keyword'], Context_switching_8p_64K['downrow'])[7]
print("Context_switching__8p_64K['value'] = %s\n" %(Context_switching_8p_64K['value']))

Context_switching_16p_16K = {'name':'系统进程切换时延 16p/16K ctxsw',
        'explain':'每个进程size为16K（执行任务）进程数为16上下文切换耗时',
        'keyword':'Context switching',
        'downrow':5,
        }
Context_switching_16p_16K['value'] = get_file_lmbench_value(Context_switching_16p_16K['keyword'], Context_switching_16p_16K['downrow'])[8]
print("Context_switching_16p_16K['value'] = %s\n" %(Context_switching_16p_16K['value']))

Context_switching_16p_64K = {'name':'系统进程切换时延 16p/64K ctxsw',
        'explain':'每个进程size为64K（执行任务）进程数为16上下文切换耗时',
        'keyword':'Context switching',
        'downrow':5,
        }
Context_switching_16p_64K['value'] = get_file_lmbench_value(Context_switching_16p_64K['keyword'], Context_switching_16p_64K['downrow'])[9]
print("Context_switching_16p_64K['value'] = %s\n" %(Context_switching_16p_64K['value']))


row += 1
write_lmbench_excel(Context_switching_2p_0k, row)
row += 1
write_lmbench_excel(Context_switching_2p_16K, row)
row += 1
write_lmbench_excel(Context_switching_2p_64K, row)
row += 1
write_lmbench_excel(Context_switching_8p_16K, row)
row += 1
write_lmbench_excel(Context_switching_8p_64K, row)
row += 1
write_lmbench_excel(Context_switching_16p_16K, row)
row += 1
write_lmbench_excel(Context_switching_16p_64K, row)


row += 1
title13 = "*Local* Communication latencies in microseconds - smaller is better     本地通讯延时"
# 在表格的第一行第一列插入变量名为title值
ws.cell(row=row, column=1, value=title13)
row += 1
title14 = "单位：μs"
# 在表格的第二行第一列插入变量名为title值
ws.cell(row=row, column=1, value=title14)
Local_Communication_Pipe = {'name':'进程pipe通信时延 Pipe',
        'explain':'两个没有具体任务的进程用unix pipe通信，一个token在两个进程间来回传递，传递一个来回所耗用的平均时间',
        'keyword':'Communication latencies in',
        'downrow':5,
        }
Local_Communication_Pipe['value'] = get_file_lmbench_value(Local_Communication_Pipe['keyword'], Local_Communication_Pipe['downrow'])[4]
print("Local_Communication_Pipe['value'] = %s\n" %(Local_Communication_Pipe['value']))

Local_Communication_AF_UNIX = {'name':'进程unix socket通信时延 AF UNIX',
        'explain':'两个没有具体任务的进程用unix socket通信，一个token在两个进程间来回传递，传递一个来回所耗用的平均时间',
        'keyword':'Communication latencies in',
        'downrow':5,
        }
Local_Communication_AF_UNIX['value'] = get_file_lmbench_value(Local_Communication_AF_UNIX['keyword'], Local_Communication_AF_UNIX['downrow'])[5]
print("Local_Communication_AF_UNIX['value'] = %s\n" %(Local_Communication_AF_UNIX['value']))

Local_Communication_UDP = {'name':'进程udp通信时延 UDP',
        'explain':'两个没有具体任务的进程用UDP/IP 通信，一个token在两个进程间来回传递，传递一个来回所耗用的平均时间',
        'keyword':'Communication latencies in',
        'downrow':5,
        }
Local_Communication_UDP['value'] = get_file_lmbench_value(Local_Communication_UDP['keyword'], Local_Communication_UDP['downrow'])[6]
print("Local_Communication_UDP['value'] = %s\n" %(Local_Communication_UDP['value']))


Local_Communication_TCP = {'name':'进程tcp通信时延 TCP',
        'explain':'两个没有具体任务的进程用TCP/IP 通信，一个token在两个进程间来回传递，传递一个来回所耗用的平均时间',
        'keyword':'Communication latencies in',
        'downrow':5,
        }
Local_Communication_TCP['value'] = get_file_lmbench_value(Local_Communication_TCP['keyword'], Local_Communication_TCP['downrow'])[8]
print("Local_Communication_TCP['value'] = %s\n" %(Local_Communication_TCP['value']))

Local_Communication_TCP_conn = {'name':'进程tcp建立连接时延 TCP conn',
        'explain':'创建一个AF_INET (aka TCP/IP) socket，并连接到远程主机所耗用的时间，这个时间仅指创建socket和建立连接本身，不包括解析主机名等等其他动作所用时间。',
        'keyword':'Communication latencies in',
        'downrow':5,
        }
Local_Communication_TCP_conn['value'] = get_file_lmbench_value(Local_Communication_TCP_conn['keyword'], Local_Communication_TCP_conn['downrow'])[10]
print("Local_Communication_TCP_conn['value'] = %s\n" %(Local_Communication_TCP_conn['value']))
row += 1
write_lmbench_excel(Local_Communication_Pipe, row)
row += 1
write_lmbench_excel(Local_Communication_AF_UNIX, row)
row += 1
write_lmbench_excel(Local_Communication_UDP, row)
row += 1
write_lmbench_excel(Local_Communication_TCP, row)
row += 1
write_lmbench_excel(Local_Communication_TCP_conn, row)

row += 1
title15 = "File & VM system latencies in microseconds - smaller is better      文档、内存延时"
# 在表格的第一行第一列插入变量名为title值
ws.cell(row=row, column=1, value=title15)
row += 1
title16 = "单位：μs"
# 在表格的第二行第一列插入变量名为title值
ws.cell(row=row, column=1, value=title16)
VM_system_latencies_Mmap_Latency = {'name':'文件mmap映射延时Tmmap Mmap Latency',
        'explain':'将指定文件的开头n个字节map到内存，然后umap，并记录每次map和umap共耗用的时间；记录的是每次耗用时间的最大值',
        'keyword':'VM system latencies',
        'downrow':5,
        }
VM_system_latencies_Mmap_Latency['value'] = ' '
print("VM_system_latencies_Mmap_Latency['value'] = %s\n" %(VM_system_latencies_Mmap_Latency['value']))

VM_system_latencies_Prot_Fault = {'name':'页保护延时 Prot Fault',
        'explain':'保护页延时时间',
        'keyword':'VM system latencies',
        'downrow':5,
        }
VM_system_latencies_Prot_Fault['value'] = get_file_lmbench_value(VM_system_latencies_Prot_Fault['keyword'], VM_system_latencies_Prot_Fault ['downrow'])[7]
print("VM_system_latencies_Prot_Fault['value'] = %s\n" %(VM_system_latencies_Prot_Fault['value']))

VM_system_latencies_Page_Fault = {'name':'缺页异常延时  Page Fault',
        'explain':'缺页延时时间',
        'keyword':'VM system latencies',
        'downrow':5,
        }
VM_system_latencies_Page_Fault['value'] = ' '
print("VM_system_latencies_Page_Fault['value'] = %s\n" %(VM_system_latencies_Page_Fault['value']))

VM_system_latencies_100fd_selct = {'name':'100fd selct',
        'explain':' 对100个文档描述符配置select的时间',
        'keyword':'VM system latencies',
        'downrow':5,
        }
VM_system_latencies_100fd_selct['value'] = get_file_lmbench_value(VM_system_latencies_100fd_selct['keyword'], VM_system_latencies_100fd_selct['downrow'])[8]
print("VM_system_latencies_100fd_selct['value'] = %s\n" %(VM_system_latencies_100fd_selct['value']))

row += 1
write_lmbench_excel(VM_system_latencies_Mmap_Latency, row)
row += 1
write_lmbench_excel(VM_system_latencies_Prot_Fault, row)
row += 1
write_lmbench_excel(VM_system_latencies_Page_Fault, row)
row += 1
write_lmbench_excel(VM_system_latencies_100fd_selct, row)

row += 1
title17 = "*Local* Communication bandwidths in MB/s - bigger is better    本地通信带宽"
# 在表格的第一行第一列插入变量名为title值
ws.cell(row=row, column=1, value=title17)
row += 1
title18 = "单位：MB/S"
# 在表格的第二行第一列插入变量名为title值
ws.cell(row=row, column=1, value=title18)
Communication_bandwidths_Pipe = {'name':'进程pipe通信带宽 Pipe',
        'explain':'两个进程间建立一个unix pipe，pipe的每个chunk为64K，通过该管道移动50M数据所用的时间',
        'keyword':'Communication bandwidths in',
        'downrow':5,
        }
Communication_bandwidths_Pipe['value'] = get_file_lmbench_value(Communication_bandwidths_Pipe['keyword'], Communication_bandwidths_Pipe['downrow'])[3]
print("Communication_bandwidths_Pipe['value'] = %s\n" %(Communication_bandwidths_Pipe['value']))

Communication_bandwidths_AF_UNIX = {'name':'进程unix socket通信带宽  AF UNIX',
        'explain':'两个进程间建立一个unix stream socket，每个chunk为64K，通过该socket移动10M数据所用的时间',
        'keyword':'Communication bandwidths in',
        'downrow':5,
        }
Communication_bandwidths_AF_UNIX['value'] = get_file_lmbench_value(Communication_bandwidths_AF_UNIX['keyword'], Communication_bandwidths_AF_UNIX['downrow'])[4]
print("Communication_bandwidths_AF_UNIX['value'] = %s\n" %(Communication_bandwidths_AF_UNIX['value']))


Communication_bandwidths_TCP = {'name':'进程tcp通信带宽 TCP',
        'explain':'同Pipe，不同的是进程间通过TCP/IP socket 通信，传输的数据为3MB',
        'keyword':'Communication bandwidths in',
        'downrow':5,
        }
Communication_bandwidths_TCP['value'] = get_file_lmbench_value(Communication_bandwidths_TCP['keyword'], Communication_bandwidths_TCP['downrow'])[5]
print("Communication_bandwidths_TCP['value'] = %s\n" %(Communication_bandwidths_TCP['value']))

Communication_bandwidths_File_reread = {'name':'文件读取带宽 File reread',
        'explain':'读文件并把他们汇总起来所用的时间',
        'keyword':'Communication bandwidths in',
        'downrow':5,
        }
Communication_bandwidths_File_reread['value'] = get_file_lmbench_value(Communication_bandwidths_File_reread['keyword'], Communication_bandwidths_File_reread['downrow'])[6]
print("Communication_bandwidths_File_reread['value'] = %s\n" %(Communication_bandwidths_File_reread['value']))

Communication_bandwidths_Mmap_reread = {'name':'文件内存映射带宽 Mmap reread',
        'explain':'将文件map到内存中，从内存中读文件并把他们汇总起来所用的时间',
        'keyword':'Communication bandwidths in',
        'downrow':5,
        }
Communication_bandwidths_Mmap_reread['value'] = get_file_lmbench_value(Communication_bandwidths_Mmap_reread['keyword'], Communication_bandwidths_Mmap_reread['downrow'])[7]
print("Communication_bandwidths_Mmap_reread['value'] = %s\n" %(Communication_bandwidths_Mmap_reread['value']))

Communication_bandwidths_Bcopy_libc = {'name':'内存memcpy拷贝带宽 Bcopy libc',
        'explain':'从指定内存区域拷贝指定数目的字节内容到指定的另一个内存区域的速度',
        'keyword':'Communication bandwidths in',
        'downrow':5,
        }
Communication_bandwidths_Bcopy_libc['value'] = get_file_lmbench_value(Communication_bandwidths_Bcopy_libc['keyword'], Communication_bandwidths_Bcopy_libc['downrow'])[8]
print("Communication_bandwidths_Bcopy_libc['value'] = %s\n" %(Communication_bandwidths_Bcopy_libc['value']))

Communication_bandwidths_Bcopy_hand = {'name':'内存赋值拷贝带宽 Bcopy hand',
        'explain':'把数据从磁盘上一个位置拷贝到另一个位置所用的时间',
        'keyword':'Communication bandwidths in',
        'downrow':5,
        }
Communication_bandwidths_Bcopy_hand['value'] = get_file_lmbench_value(Communication_bandwidths_Bcopy_hand['keyword'], Communication_bandwidths_Bcopy_hand['downrow'])[9]
print("Communication_bandwidths_Bcopy_hand['value'] = %s\n" %(Communication_bandwidths_Bcopy_hand['value']))

Communication_bandwidths_Mem_read = {'name':'内存读取累加带宽 Mem read',
        'explain':'累加数组中的整数值，测试把数据读入processor的带宽',
        'keyword':'Communication bandwidths in',
        'downrow':5,
        }
Communication_bandwidths_Mem_read['value'] = get_file_lmbench_value(Communication_bandwidths_Mem_read['keyword'], Communication_bandwidths_Mem_read['downrow'])[10]
print("Communication_bandwidths_Mem_read['value'] = %s\n" %(Communication_bandwidths_Mem_read['value']))
Communication_bandwidths_Mem_write = {'name':'内存写入带宽 Mem write',
        'explain':'把整数数组的每个成员设置为1，测试写数据到内存的带宽',
        'keyword':'Communication bandwidths in',
        'downrow':5,
        }
Communication_bandwidths_Mem_write['value'] = get_file_lmbench_value(Communication_bandwidths_Mem_write['keyword'], Communication_bandwidths_Mem_write['downrow'])[11]
print("Communication_bandwidths_Mem_write['value'] = %s\n" %(Communication_bandwidths_Mem_write['value']))


row += 1
write_lmbench_excel(Communication_bandwidths_Pipe, row)
row += 1
write_lmbench_excel(Communication_bandwidths_AF_UNIX, row)
row += 1
write_lmbench_excel(Communication_bandwidths_TCP, row)
row += 1
write_lmbench_excel(Communication_bandwidths_File_reread, row)
row += 1
write_lmbench_excel(Communication_bandwidths_Mmap_reread, row)
row += 1
write_lmbench_excel(Communication_bandwidths_Bcopy_libc, row)
row += 1
write_lmbench_excel(Communication_bandwidths_Bcopy_hand, row)
row += 1
write_lmbench_excel(Communication_bandwidths_Mem_read, row)
row += 1
write_lmbench_excel(Communication_bandwidths_Mem_write, row)



# # 保存表格
bw.save('lmbench3.xlsx')
