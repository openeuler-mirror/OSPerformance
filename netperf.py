# -*- coding: utf-8 -*-
# /usr/bin/python
#########################################################
# Function :cangkujiance                                #
# Platform :UnionTech OS                                #
# Version  :1.0                                         #
# Date     :2022-08-29                                  #
# Author   :hanshuang                                   #
# Contact  :hanshuang@uniontech.com                     #
# Company  :UnionTech                                   #
#########################################################
# 5.6 网络性能
# UDP_STREAM 需要在同一网段并且服务端和客户端选择加密必须都加密，选择手动配置都手动配置。
# 导入openpyxl库
import os
import openpyxl
from openpyxl.styles import Alignment
import linecache
import re

path = './'
print(os.getcwd())
def get_file_netperf_value(File_name, keyword, downrow):
    with open(File_name, "r+") as f:
        readlines = f.readlines()
        number = 1
        for readline in readlines:
            if readline.find(keyword) != -1:
                break
            else:
                number += 1

    row = number + downrow
    read_specified_line_value = linecache.getline(File_name, row)
    netperf_value_tmp = read_specified_line_value.split()
    netperf_value = netperf_value_tmp[len(netperf_value_tmp)-1]
    return netperf_value

def write_netperf_excel(dict, row_1):
    ws.cell(row=row_1, column=1, value=dict['name'])
    ws.cell(row=row_1, column=2, value=dict['explain'])
    ws.cell(row=row_1, column=3, value=dict['netperf_value'])

with open(path + 'netperf.conf', "r+") as f:
        netperf_conf_readlines = f.readlines()
        for netperf_conf_readline in netperf_conf_readlines:
            if netperf_conf_readline.find("SERVERIP") != -1:
                SERVERIP = netperf_conf_readline.split("=")[1]
                SERVERIP = SERVERIP[:-1]
                print("%s\n" %(SERVERIP))
            if netperf_conf_readline.find("RUNTIME") != -1:
                RUNTIME = netperf_conf_readline.split("=")[1]
                RUNTIME = RUNTIME[:-1]
                print("%s\n" %(RUNTIME))
            if netperf_conf_readline.find("NETPERF_SERVER_PASSPW") != -1:
                NETPERF_SERVER_PASSPW = netperf_conf_readline.split("=")[1]
                NETPERF_SERVER_PASSPW = NETPERF_SERVER_PASSPW[:-1]
                print("%s\n" %(NETPERF_SERVER_PASSPW))
        f.close()

# os.system("sshpass -p '%s' ssh root@%s 'mkdir /root/src'" %(NETPERF_SERVER_PASSPW, SERVERIP))
# os.system("sshpass -p '%s' scp src/fio-2.1.10.tar.gz  root@%s:/root/src" %(NETPERF_SERVER_PASSPW, SERVERIP))
# os.system("sshpass -p '%s' ssh root@%s 'bash -s' < netperf_server.sh" %(NETPERF_SERVER_PASSPW, SERVERIP))

os.system("sh netperf_client.sh %s %s" %(SERVERIP, RUNTIME))

# 新建一个工作薄
bw = openpyxl.Workbook()
# 激活表格
ws = bw.active
# 新建一个名称为netperf_sheet1的表单
ws.title = 'netperf_sheet1'
title = "netperf (数据越大性能越好)"
# 在表格的第一行第一列插入变量名为title值
ws.cell(row=1, column=1, value=title)
TCP_STREAM = {'name':'TCP 批量传输 (MB/s)', 
        'explain':'TCP 批量数据传输的吞吐量', 
        'keyword':'Throughput',
        'downrow':3,
        'file_name':path + 'report/netperf/TCP_STREAM.txt'
        }   
TCP_STREAM['netperf_value'] = get_file_netperf_value(TCP_STREAM['file_name'],TCP_STREAM['keyword'], TCP_STREAM['downrow'])
row = 2
write_netperf_excel(TCP_STREAM, row)

UDP_STREAM = {'name':'UDP 批量传输 (MB/s)',
        'explain':'UDP 批量数据传输的吞吐量',
        'keyword':'Throughput',
        'downrow':3,
        'file_name':path + 'report/netperf/UDP_STREAM.txt'
        }
UDP_STREAM['netperf_value'] = get_file_netperf_value(UDP_STREAM['file_name'], UDP_STREAM['keyword'], UDP_STREAM['downrow'])
row += 1
write_netperf_excel(UDP_STREAM, row)

TCP_RR = {'name':'TCP 请求响应速度 (次/s)',
        'explain':'TCP 长连接请求应答的平均交易率',
        'keyword':'Trans.',
        'downrow':4,
        'file_name':path + 'report/netperf/TCP_RR.txt'
        }
TCP_RR['netperf_value'] = get_file_netperf_value(TCP_RR['file_name'], TCP_RR['keyword'], TCP_RR['downrow'])
row += 1
write_netperf_excel(TCP_RR, row)
 
TCP_CRR = {'name':'HTTP应用测试 (次/s)',
        'explain':'TCP 短连接请求应答的平均交易率',
        'keyword':'Trans.',
        'downrow':4,
        'file_name':path + 'report/netperf/TCP_CRR.txt'
        }
TCP_CRR['netperf_value'] = get_file_netperf_value(TCP_CRR['file_name'], TCP_CRR['keyword'], TCP_CRR['downrow'])
row += 1
write_netperf_excel(TCP_CRR, row)


UDP_RR = {'name':'UDP 请求响应速度 (次/s)',
        'explain':'UDP 分组进行请求应答的平均交易率',
        'keyword':'Trans.',
        'downrow':4,
        'file_name':path + 'report/netperf/UDP_RR.txt'
        }
UDP_RR['netperf_value'] = get_file_netperf_value(UDP_RR['file_name'], UDP_RR['keyword'], UDP_RR['downrow'])
row += 1
write_netperf_excel(UDP_RR, row)

omni = {'name':'网络响应时间 (μs)',
        'explain':'测试服务端和客户端的平均时延',
        'keyword':'Mean',
        'downrow':4,
        'file_name':path + 'report/netperf/omni.txt'
        }
omni['netperf_value'] = get_file_netperf_value(omni['file_name'], omni['keyword'], omni['downrow'])
row += 1
write_netperf_excel(omni, row)


# # 保存表格
bw.save(path+'report/netperf/netperf.xlsx')
