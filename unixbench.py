import os
import xlsxwriter
import linecache
import os
import openpyxl
from openpyxl.styles import Alignment
import linecache
import re
import datetime
import socket

path = './'

print(os.getcwd())
#os.system('sh unixbench2.sh')


#"""
now=datetime.datetime.now()
date=now.strftime("%Y-%m-%d")
print(date)

hostname=socket.gethostname()
print(hostname)

file1=path+'report/unixbench_results/'+hostname+'-'+date+'-01'
print(file1)

def get_unixbench_value(keyword, downrow, file):
    with open(file, "r+") as f:
        readlines = f.readlines()
        lenth = len(readlines)
        print(lenth)
        number = 1
        for readline in readlines:
            if readline.find(keyword) != -1:
                break
            else:
                number += 1

#    print("number = %d" % number)
    row = number + downrow
    read_value = linecache.getline(file, row)
    value = read_value.split()
#    print("value = %s\n" % value)
    return value


def write_unixbench_excel(dict, row_1):
    ws.cell(row=row_1, column=1, value=dict['name'])
    ws.cell(row=row_1, column=2, value=dict['explain'])
    ws.cell(row=row_1, column=3, value=dict['value'])


if os.path.exists(path + 'report/' + 'unixbench.xlsx'):
    os.system('rm -rf ./report/unixbench.xlsx')

# 新建一个工作薄
bw = openpyxl.Workbook()
# 激活表格
ws = bw.active
# 新建一个名称为lmbench_sheet1的表单
ws.title = 'unixbench_sheet1'
heading = ["测试项", "测试说明", "测试结果"]
test_item = ["单核测试", "满核测试"]
# 在表格的第一行第一列插入变量名为title值
ws.cell(row=1, column=1, value=heading[0])
ws.cell(row=1, column=2, value=heading[1])
ws.cell(row=1, column=3, value=heading[2])

ws.cell(row=2, column=1, value=test_item[0])

Dhrystone_2_using_register_variables = {'name': 'Dhrystone_2_using_register_variables',
                                        'explain': '计算和比较计算机性能',
                                        'keyword': '1 parallel',
                                        'downrow': 16,
                                        }
Dhrystone_2_using_register_variables['value'] = get_unixbench_value(Dhrystone_2_using_register_variables['keyword'],
                                                                    Dhrystone_2_using_register_variables['downrow'],file1)[-1]

Double_Precision_Whetstone = {'name': 'Dhrystone_2_using_register_variables',
                              'explain': '测量浮点运算速度和效率',
                              'keyword': '1 parallel',
                              'downrow': 17,
                              }
Double_Precision_Whetstone['value'] = get_unixbench_value(Double_Precision_Whetstone['keyword'],
                                                          Double_Precision_Whetstone['downrow'],file1)[-1]

Execl_Throughput = {'name': 'Execl Throughput',
                    'explain': '计算每秒钟出现的execl调用数',
                    'keyword': '1 parallel',
                    'downrow': 18,
                    }
Execl_Throughput['value'] = get_unixbench_value(Execl_Throughput['keyword'], Execl_Throughput['downrow'],file1)[-1]

File_Copy_1024_bufsize_2000 = {'name': 'File_Copy_1024_bufsize_2000_maxblocks',
                               'explain': '获得在指定时间内能够读、写、复制的字符数目-1',
                               'keyword': '1 parallel',
                               'downrow': 19,
                               }
File_Copy_1024_bufsize_2000['value'] = get_unixbench_value(File_Copy_1024_bufsize_2000['keyword'],
                                                           File_Copy_1024_bufsize_2000['downrow'],file1)[-1]

File_Copy_256_bufsize_500 = {'name': 'File_Copy_256_bufsize_500_maxblocks',
                             'explain': '获得在指定时间内能够读、写、复制的字符数目-2',
                             'keyword': '1 parallel',
                             'downrow': 20,
                             }
File_Copy_256_bufsize_500['value'] = get_unixbench_value(File_Copy_256_bufsize_500['keyword'],
                                                         File_Copy_256_bufsize_500['downrow'],file1)[-1]

File_Copy_4096_bufsize_8000 = {'name': 'File_Copy_4096_bufsize_8000_maxblocks',
                               'explain': '获得在指定时间内能够读、写、复制的字符数目-3',
                               'keyword': '1 parallel',
                               'downrow': 21,
                               }
File_Copy_4096_bufsize_8000['value'] = get_unixbench_value(File_Copy_4096_bufsize_8000['keyword'],
                                                           File_Copy_4096_bufsize_8000['downrow'],file1)[-1]

Pipe_Throughput = {'name': 'Pipe Throughput',
                   'explain': '管道吞吐量',
                   'keyword': '1 parallel',
                   'downrow': 22,
                   }
Pipe_Throughput['value'] = get_unixbench_value(Pipe_Throughput['keyword'], Pipe_Throughput['downrow'],file1)[-1]

Pipe_based_Context_Switching = {'name': 'Pipe_based_Context_Switching',
                                'explain': '计算两个进程通过管道交换一个增长的整数的次数',
                                'keyword': '1 parallel',
                                'downrow': 23,
                                }
Pipe_based_Context_Switching['value'] = get_unixbench_value(Pipe_based_Context_Switching['keyword'],
                                                            Pipe_based_Context_Switching['downrow'],file1)[-1]

Process_Creation = {'name': 'Process_Creation',
                    'explain': '计算一个进程派生和收获一个马上退出的子进程的次数',
                    'keyword': '1 parallel',
                    'downrow': 24,
                    }
Process_Creation['value'] = get_unixbench_value(Process_Creation['keyword'], Process_Creation['downrow'],file1)[-1]

Shell_Scripts_1 = {'name': 'Shell_Scripts(1 concurrent)',
                   'explain': '每秒进程能够启动和收获一组1个shell脚本程序的并行的拷贝的次数',
                   'keyword': '1 parallel',
                   'downrow': 25,
                   }
Shell_Scripts_1['value'] = get_unixbench_value(Shell_Scripts_1['keyword'], Shell_Scripts_1['downrow'],file1)[-1]

Shell_Scripts_8 = {'name': 'Process_Creation',
                   'explain': '每秒进程能够启动和收获一组8个shell脚本程序的并行的拷贝的次数',
                   'keyword': '1 parallel',
                   'downrow': 26,
                   }
Shell_Scripts_8['value'] = get_unixbench_value(Shell_Scripts_8['keyword'], Shell_Scripts_8['downrow'],file1)[-1]

System_Call_Overhead = {'name': 'System Call Overhead',
                        'explain': '进入和离开系统内核的消耗',
                        'keyword': '1 parallel',
                        'downrow': 27,
                        }
System_Call_Overhead['value'] = get_unixbench_value(System_Call_Overhead['keyword'], System_Call_Overhead['downrow'],file1)[
    -1]

System_Benchmarks_Index_Score = {'name': 'System Benchmarks Index Score',
                                 'explain': '系统基准指数得分',
                                 'keyword': '1 parallel',
                                 'downrow': 29,
                                 }
System_Benchmarks_Index_Score['value'] = get_unixbench_value(System_Benchmarks_Index_Score['keyword'],
                                                             System_Benchmarks_Index_Score['downrow'],file1)[-1]

row = 3
write_unixbench_excel(Dhrystone_2_using_register_variables, row)
row += 1
write_unixbench_excel(Double_Precision_Whetstone, row)
row += 1
write_unixbench_excel(Execl_Throughput, row)
row += 1
write_unixbench_excel(File_Copy_1024_bufsize_2000, row)
row += 1
write_unixbench_excel(File_Copy_256_bufsize_500, row)
row += 1
write_unixbench_excel(File_Copy_4096_bufsize_8000, row)
row += 1
write_unixbench_excel(Pipe_Throughput, row)
row += 1
write_unixbench_excel(Pipe_based_Context_Switching, row)
row += 1
write_unixbench_excel(Process_Creation, row)
row += 1
write_unixbench_excel(Shell_Scripts_1, row)
row += 1
write_unixbench_excel(Shell_Scripts_8, row)
row += 1
write_unixbench_excel(System_Call_Overhead, row)
row += 1
write_unixbench_excel(System_Benchmarks_Index_Score, row)
print(row)

# 满核测试
#"""
from multiprocessing import cpu_count
print(cpu_count())
cpus=cpu_count()
keyword1=str(cpus)+' '+'parallel'
print(keyword1)
#"""

ws.cell(row=16, column=1, value=test_item[1])

Dhrystone_2using_register_variables1 = {'name': 'Dhrystone_2_using_register_variables',
                                        'explain': '计算和比较计算机性能',
                                        'downrow': 16,
                                        }
Dhrystone_2using_register_variables1['value'] = get_unixbench_value(keyword1,
                                                                    Dhrystone_2using_register_variables1['downrow'],file1)[-1]

print(file1)

Double_Precision_Whetstone1 = {'name': 'Dhrystone_2_using_register_variables',
                               'explain': '测量浮点运算速度和效率',
                               'downrow': 17,
                               }
Double_Precision_Whetstone1['value'] = get_unixbench_value(keyword1,
                                                           Double_Precision_Whetstone1['downrow'],file1)[-1]

Execl_Throughput1 = {'name': 'Execl Throughput',
                     'explain': '计算每秒钟出现的execl调用数',
                     'downrow': 18,
                     }
Execl_Throughput1['value'] = get_unixbench_value(keyword1, Execl_Throughput1['downrow'],file1)[-1]

File_Copy_1024_bufsize_2000_1 = {'name': 'File_Copy_1024_bufsize_2000_maxblocks',
                                 'explain': '获得在指定时间内能够读、写、复制的字符数目-1',
                                 'downrow': 19,
                                 }
File_Copy_1024_bufsize_2000_1['value'] = get_unixbench_value(keyword1,
                                                             File_Copy_1024_bufsize_2000_1['downrow'],file1)[-1]

File_Copy_256_bufsize_500_1 = {'name': 'File_Copy_256_bufsize_500_maxblocks',
                               'explain': '获得在指定时间内能够读、写、复制的字符数目-2',
                               'downrow': 20,
                               }
File_Copy_256_bufsize_500_1['value'] = get_unixbench_value(keyword1,
                                                           File_Copy_256_bufsize_500_1['downrow'],file1)[-1]

File_Copy_4096_bufsize_8000_1 = {'name': 'File_Copy_4096_bufsize_8000_maxblocks',
                                 'explain': '获得在指定时间内能够读、写、复制的字符数目-3',
                                 'downrow': 21,
                                 }
File_Copy_4096_bufsize_8000_1['value'] = get_unixbench_value(keyword1,
                                                             File_Copy_4096_bufsize_8000_1['downrow'],file1)[-1]

Pipe_Throughput1 = {'name': 'Pipe Throughput',
                    'explain': '管道吞吐量',
                    'downrow': 22,
                    }
Pipe_Throughput1['value'] = get_unixbench_value(keyword1, Pipe_Throughput1['downrow'],file1)[-1]

Pipe_based_Context_Switching1 = {'name': 'Pipe_based_Context_Switching',
                                 'explain': '计算两个进程通过管道交换一个增长的整数的次数',
                                 'downrow': 23,
                                 }
Pipe_based_Context_Switching1['value'] = get_unixbench_value(keyword1,
                                                             Pipe_based_Context_Switching1['downrow'],file1)[-1]

Process_Creation1 = {'name': 'Process_Creation',
                     'explain': '计算一个进程派生和收获一个马上退出的子进程的次数',
                     'downrow': 24,
                     }
Process_Creation1['value'] = get_unixbench_value(keyword1, Process_Creation1['downrow'],file1)[-1]

Shell_Scripts_11 = {'name': 'Shell_Scripts(1 concurrent)',
                    'explain': '每秒进程能够启动和收获一组1个shell脚本程序的并行的拷贝的次数',
                    'downrow': 25,
                    }
Shell_Scripts_11['value'] = get_unixbench_value(keyword1, Shell_Scripts_11['downrow'],file1)[-1]

Shell_Scripts_81 = {'name': 'Process_Creation',
                    'explain': '每秒进程能够启动和收获一组8个shell脚本程序的并行的拷贝的次数',
                    'downrow': 26,
                    }
Shell_Scripts_81['value'] = get_unixbench_value(keyword1, Shell_Scripts_81['downrow'],file1)[-1]

System_Call_Overhead1 = {'name': 'System Call Overhead',
                         'explain': '进入和离开系统内核的消耗',
                         'downrow': 27,
                         }
System_Call_Overhead1['value'] = get_unixbench_value(keyword1, System_Call_Overhead1['downrow'],file1)[-1]

System_Benchmarks_Index_Score1 = {'name': 'System Benchmarks Index Score',
                                  'explain': '系统基准指数得分',
                                  'downrow': 29,
                                  }
System_Benchmarks_Index_Score1['value'] = get_unixbench_value(keyword1,
                                                              System_Benchmarks_Index_Score1['downrow'],file1)[-1]

row = 17
write_unixbench_excel(Dhrystone_2using_register_variables1, row)
row += 1
write_unixbench_excel(Double_Precision_Whetstone1, row)
row += 1
write_unixbench_excel(Execl_Throughput1, row)
row += 1
write_unixbench_excel(File_Copy_1024_bufsize_2000_1, row)
row += 1
write_unixbench_excel(File_Copy_256_bufsize_500_1, row)
row += 1
write_unixbench_excel(File_Copy_4096_bufsize_8000_1, row)
row += 1
write_unixbench_excel(Pipe_Throughput1, row)
row += 1
write_unixbench_excel(Pipe_based_Context_Switching1, row)
row += 1
write_unixbench_excel(Process_Creation1, row)
row += 1
write_unixbench_excel(Shell_Scripts_11, row)
row += 1
write_unixbench_excel(Shell_Scripts_81, row)
row += 1
write_unixbench_excel(System_Call_Overhead1, row)
row += 1
write_unixbench_excel(System_Benchmarks_Index_Score1, row)
row += 1

bw.save(path + 'report/' + 'unixbench.xlsx')

