# -*- coding: utf-8 -*-
# /usr/bin/python
#########################################################
# Function :cangkujiance                                #
# Platform :UnionTech OS                                #
# Version  :1.0                                         #
# Date     :2022-08-29                                  #
# Author   :hanshuang                                   #
# Contact  :hanshuang@uniontech.com                     #
# Company  :UnionTech                                   #
#########################################################
# 5.5　文件读写性能
# 导入openpyxl库
import os
import openpyxl
from openpyxl.styles import Alignment
import linecache
import re
import statistics
import time

path = './'
print(os.getcwd())
os.system('sh fio_compile.sh')
#os.system('sh fio_test1.sh')
def get_file_bw_value(File_name):
    with open(File_name, "r+") as f:
        readlines = f.readlines()
        number = 1
        for readline in readlines:
            if readline.find('BW') != -1:
                f.close()
                break
            else:
                number += 1
    read_specified_line_value = linecache.getline(File_name, number)
    fio_bw_rate = read_specified_line_value.split(",")[1].split("=")[1].split()[0]
    print("fio_bw_rate=%s" %fio_bw_rate)
    rate = re.sub("[0-9 .]", "", fio_bw_rate)
    print("rate=%s" %rate)
    bw_value = 1.0
    if rate == "B/s":
        fio_bw = fio_bw_rate.strip('B/s')
        bw_tmp = float(fio_bw) / 1024 
        bw_value = round(float(bw_tmp), 2)
    elif rate == "KiB/s":
        fio_bw = fio_bw_rate.strip('KiB/s')
        print("fio_bw = %s" %(fio_bw))
        bw_tmp = float(fio_bw) 
        bw_value = round(bw_tmp, 2)
    elif rate == "MiB/s":
        fio_bw = fio_bw_rate.strip('MiB/s')
        bw_tmp = float(fio_bw) 
        bw_value = round(bw_tmp, 2)
        print("bw_value=%s" %bw_value)
    return bw_value

def get_file_iops_value(File_name):
    with open(File_name, "r+") as f:
        readlines = f.readlines()
        number = 1
        for readline in readlines:
            if readline.find('IOPS') != -1:
                f.close()
                break
            else:
                number += 1
    read_specified_line_value = linecache.getline(File_name, number)
    fio_iops = read_specified_line_value.split(",")[0].split("=")[1]
    iops_tmp = fio_iops.strip('k')
    fio_iops_value =float(iops_tmp)
    print("fio_iops_value=%s" %fio_iops_value)
    return fio_iops_value

def write_bw_excel(dict, row_1):
    ws.cell(row=row_1, column=1, value=dict['name'])
    ws.cell(row=row_1, column=2, value=dict['explain'])
    bw_value = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()) + "  "  + str(dict['bw_value'])
    ws.cell(row=row_1, column=3, value=bw_value)
    ws.cell(row=row_1, column=3, value=dict['bw_value'])

def write_iops_excel(dict, row_1):
    ws.cell(row=row_1, column=1, value=dict['name'])
    ws.cell(row=row_1, column=2, value=dict['explain'])
    ipos_value = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()) + "  "  + str(dict['iops_value'])
    ws.cell(row=row_1, column=3, value=ipos_value)

with open(path + 'fio.conf', "r+") as f:
        fio_conf_readlines = f.readlines()
        for fio_conf_readline in fio_conf_readlines:
            if fio_conf_readline.find("FILENAME") != -1:
                FILENAME = fio_conf_readline.split("=")[1]
                FILENAME = FILENAME[:-1]
                print("%s\n" %(FILENAME))
            if fio_conf_readline.find("RUNTIME") != -1:
                RUNTIME = fio_conf_readline.split("=")[1]
                RUNTIME = RUNTIME[:-1]
                print("%s\n" %(RUNTIME))
        f.close()


os.system("bash fio_test.sh %s %s" %(FILENAME, RUNTIME))

# 新建一个工作薄
bw = openpyxl.Workbook()
# 激活表格
ws = bw.active
# 新建一个名称为stream_sheet1的表单
ws.title = 'fio_sheet1'
title = "4k"
# 在表格的第一行第一列插入变量名为ti值
ws.cell(row=1, column=1, value=title)

# fio_test.sh参数解释
# 第一个参数为测试的硬盘比如/dev/sda
# 第二个参数为运行时间可以为数字
row = 1
os.system("bash fio_test1.sh %s %s" %(FILENAME, RUNTIME))
fio_write_4k_bw = {'name':'100%顺序写模式 bw (KB/s)\n', 
        'explain':'顺序写磁盘的吞吐量\n', 
        'file_name':path + 'report/fio_result/4k/write_4k.txt'
        }   

fio_write_4k_bw['bw_value'] = get_file_bw_value(fio_write_4k_bw['file_name'])
#print("aaaaaaaaaaaaaa111111 %d \n" %(fio_write_4k_bw['bw_value']))
row += 1
write_bw_excel(fio_write_4k_bw, row)

fio_write_4k_iops = {'name':'100%顺序写模式 iops (次)\n',
        'explain':'顺序写磁盘的每秒写次数\n',
        'file_name':path + 'report/fio_result/4k/write_4k.txt'
        }
fio_write_4k_iops['iops_value'] = get_file_iops_value(fio_write_4k_iops['file_name'])
#print("aaaaaaaaaaaaaa %d \n" %(fio_write_4k_iops['ivalue']))
row += 1
write_iops_excel(fio_write_4k_iops, row)

fio_read_4k_bw = {'name':'100%顺序读模式 bw (KB/s)\n', 
        'explain':'顺序读磁盘的吞吐量\n', 
        'file_name':path + 'report/fio_result/4k/read_4k.txt'
        }
fio_read_4k_bw['bw_value'] = get_file_bw_value(fio_read_4k_bw['file_name'])
#print("bbbbbbbbbbbb %d \n" %(fio_read_4k_bw['bw_value']))
row += 1
write_bw_excel(fio_read_4k_bw, row)

fio_read_4k_iops = {'name':'100%顺序读模式 iops (次)\n',
        'explain':'顺序读磁盘的每秒读次数\n',
        'file_name':path + 'report/fio_result/4k/read_4k.txt'
        }
fio_read_4k_iops['iops_value'] = get_file_iops_value(fio_read_4k_iops['file_name'])
row += 1
write_iops_excel(fio_read_4k_iops, row)
fio_randwrite_4k_bw = {'name':'100%随机写模式 bw (KB/s)\n',
        'explain':'随机写磁盘的吞吐量\n',
        'file_name':path + 'report/fio_result/4k/randwrite_4k.txt'
        }
fio_randwrite_4k_bw['bw_value'] = get_file_bw_value(fio_randwrite_4k_bw['file_name'])
row += 1
write_bw_excel(fio_randwrite_4k_bw, row)
fio_randwrite_4k_iops = {'name':'100%随机写模式 iops (次)\n',
        'explain':'随机写磁盘的每秒写次数\n',
        'file_name':path + 'report/fio_result/4k/randwrite_4k.txt'
        }
fio_randwrite_4k_iops['iops_value'] = get_file_iops_value(fio_randwrite_4k_iops['file_name'])
row += 1
write_iops_excel(fio_randwrite_4k_iops, row)

fio_randread_4k_bw = {'name':'100%随机读模式 bw (KB/s)\n',
        'explain':'随机读磁盘的吞吐量\n',
        'file_name':path + 'report/fio_result/4k/randread_4k.txt'
        }
fio_randread_4k_bw['bw_value'] = get_file_bw_value(fio_randread_4k_bw['file_name'])
row += 1
write_bw_excel(fio_randread_4k_bw, row)

fio_randread_4k_iops = {'name':'100%随机读模式 iops (次)\n',
        'explain':'随机读磁盘的每秒读次数\n',
        'file_name':path + 'report/fio_result/4k/randread_4k.txt'
        }
fio_randread_4k_iops['iops_value'] = get_file_iops_value(fio_randread_4k_iops['file_name'])
row += 1
write_iops_excel(fio_randread_4k_iops, row)

fio_randrw_mixwrite_70_4k_bw = {'name':'写占70%随机混合读写模式 bw (KB/s)\n',
        'explain':'随机混合读写磁盘的吞吐量\n',
        'file_name':path + 'report/fio_result/4k/randrw_mixwrite_70_4k.txt'
        }
fio_randrw_mixwrite_70_4k_bw['bw_value'] = get_file_bw_value(fio_randrw_mixwrite_70_4k_bw['file_name'])
row += 1
write_bw_excel(fio_randrw_mixwrite_70_4k_bw, row)

fio_randrw_mixwrite_70_4k_iops = {'name':'写占70%随机混合读写模式 iops (次)\n',
        'explain':'随机混合读写磁盘的每秒读写次数\n',
        'file_name':path + 'report/fio_result/4k/randrw_mixwrite_70_4k.txt'
        }
fio_randrw_mixwrite_70_4k_iops['iops_value'] = get_file_iops_value(fio_randrw_mixwrite_70_4k_iops['file_name'])
row += 1
write_iops_excel(fio_randrw_mixwrite_70_4k_iops, row)

fio_randrw_mixread_70_4k_bw = {'name':'读占70%随机混合读写模式 bw (KB/s)\n',
        'explain':'随机混合读写磁盘的吞吐量\n',
        'file_name':path + 'report/fio_result/4k/randrw_mixread_70_4k.txt'
        }
fio_randrw_mixread_70_4k_bw['bw_value'] = get_file_bw_value(fio_randrw_mixread_70_4k_bw['file_name'])
row += 1
write_bw_excel(fio_randrw_mixread_70_4k_bw, row)

fio_randrw_mixread_70_4k_iops = {'name':'读占70%随机混合读写模式 iops (次)\n',
        'explain':'随机混合读写磁盘的每秒读写次数\n',
        'file_name':path + 'report/fio_result/4k/randrw_mixread_70_4k.txt'
        }
fio_randrw_mixread_70_4k_iops['iops_value'] = get_file_iops_value(fio_randrw_mixread_70_4k_iops['file_name'])
row += 1
write_iops_excel(fio_randrw_mixread_70_4k_iops, row)

bw.save('fio.xlsx')

#os.system("bash fio_test1.sh %s %s" %(FILENAME, RUNTIME))
row += 1
title1 = "128k"
# 在表格的下一行插入title为128k
row_128k = row 
ws.cell(row=row_128k, column=1, value=title1)

fio_write_128k_bw = {'name':'100%顺序写模式 bw (KB/s)\n', 
        'explain':'顺序写磁盘的吞吐量\n', 
        'file_name':path + 'report/fio_result/128k/write_128k.txt'
        }   
fio_write_128k_bw['bw_value'] = get_file_bw_value(fio_write_128k_bw['file_name'])
row += 1
write_bw_excel(fio_write_128k_bw, row)

fio_write_128k_iops = {'name':'100%顺序写模式 iops (次)\n',
        'explain':'顺序写磁盘的每秒写次数\n',
        'file_name':path + 'report/fio_result/128k/write_128k.txt'
        }
fio_write_128k_iops['iops_value'] = get_file_iops_value(fio_write_128k_iops['file_name'])
row += 1
write_iops_excel(fio_write_128k_iops, row)

fio_read_128k_bw = {'name':'100%顺序读模式 bw (KB/s)\n', 
        'explain':'顺序读磁盘的吞吐量\n', 
        'file_name':path + 'report/fio_result/128k/read_128k.txt'
        }
fio_read_128k_bw['bw_value'] = get_file_bw_value(fio_read_128k_bw['file_name'])
row += 1
write_bw_excel(fio_read_128k_bw, row)

fio_read_128k_iops = {'name':'100%顺序读模式 iops (次)\n',
        'explain':'顺序读磁盘的每秒读次数\n',
        'file_name':path + 'report/fio_result/128k/read_128k.txt'
        }
fio_read_128k_iops['iops_value'] = get_file_iops_value(fio_read_128k_iops['file_name'])
row += 1
write_iops_excel(fio_read_128k_iops, row)

fio_randwrite_128k_bw = {'name':'100%随机写模式 bw (KB/s)\n',
        'explain':'随机写磁盘的吞吐量\n',
        'file_name':path + 'report/fio_result/128k/randwrite_128k.txt'
        }
fio_randwrite_128k_bw['bw_value'] = get_file_bw_value(fio_randwrite_128k_bw['file_name'])
row += 1
write_bw_excel(fio_randwrite_128k_bw, row)

fio_randwrite_128k_iops = {'name':'100%随机写模式 iops (次)\n',
        'explain':'随机写磁盘的每秒写次数\n',
        'file_name':path + 'report/fio_result/128k/randwrite_128k.txt'
        }
fio_randwrite_128k_iops['iops_value'] = get_file_iops_value(fio_randwrite_128k_iops['file_name'])
row += 1
write_iops_excel(fio_randwrite_128k_iops, row)

fio_randread_128k_bw = {'name':'100%随机读模式 bw (KB/s)\n',
        'explain':'随机读磁i盘的吞吐量\n',
        'file_name':path + '/report/fio_result/128k/randread_128k.txt'
        }
fio_randread_128k_bw['bw_value'] = get_file_bw_value(fio_randread_128k_bw['file_name'])
row += 1
write_bw_excel(fio_randread_128k_bw, row)

fio_randread_128k_iops = {'name':'100%随机读模式 iops (次)\n',
        'explain':'随机读磁盘的每秒读次数\n',
        'file_name':path + 'report/fio_result/128k/randread_128k.txt'
        }
fio_randread_128k_iops['iops_value'] = get_file_iops_value(fio_randread_128k_iops['file_name'])
row += 1
write_iops_excel(fio_randread_128k_iops, row)

fio_randrw_mixwrite_70_128k_bw = {'name':'写占70%随机混合读写模式 bw (KB/s)\n',
        'explain':'随机混合读写磁盘的吞吐量\n',
        'file_name':path + 'report/fio_result/128k/randrw_mixwrite_70_128k.txt'
        }
fio_randrw_mixwrite_70_128k_bw['bw_value'] = get_file_bw_value(fio_randrw_mixwrite_70_128k_bw['file_name'])
row += 1
write_bw_excel(fio_randrw_mixwrite_70_128k_bw, row)

fio_randrw_mixwrite_70_128k_iops = {'name':'写占70%随机混合读写模式 iops (次)\n',
        'explain':'随机混合读写磁盘的每秒读写次数\n',
        'file_name':path + 'report/fio_result/128k/randrw_mixwrite_70_128k.txt'
        }
fio_randrw_mixwrite_70_128k_iops['iops_value'] = get_file_iops_value(fio_randrw_mixwrite_70_128k_iops['file_name'])
row += 1
write_iops_excel(fio_randrw_mixwrite_70_128k_iops, row)

fio_randrw_mixread_70_128k_bw = {'name':'读占70%随机混合读写模式 bw (KB/s)\n',
        'explain':'随机混合读写磁盘的吞吐量\n',
        'file_name':path + 'report/fio_result/128k/randrw_mixread_70_128k.txt'
        }
fio_randrw_mixread_70_128k_bw['bw_value'] = get_file_bw_value(fio_randrw_mixread_70_128k_bw['file_name'])
row += 1
write_bw_excel(fio_randrw_mixread_70_128k_bw, row)

fio_randrw_mixread_70_128k_iops = {'name':'读占70%随机混合读写模式 iops (次)\n',
        'explain':'随机混合读写磁盘的每秒读写次数\n',
        'file_name':path + 'report/fio_result/128k/randrw_mixread_70_128k.txt'
        }
fio_randrw_mixread_70_128k_iops['iops_value'] = get_file_iops_value(fio_randrw_mixread_70_128k_iops['file_name'])
row += 1
write_iops_excel(fio_randrw_mixread_70_128k_iops, row)
bw.save('fio.xlsx')
print("测试fio 128k 数据块")

#os.system("bash fio_test1.sh %s %s" %(FILENAME, RUNTIME))
title2 = "1M"
row += 1
row_1M = row
# 在表格的下一行插入title为1M
ws.cell(row=row_1M, column=1, value=title2)
fio_write_1M_bw = {'name':'100%顺序写模式 bw (KB/s)\n',
        'explain':'顺序写磁盘的吞吐量\n',
        'file_name':path + 'report/fio_result/1M/write_1M.txt'
        }
fio_write_1M_bw['bw_value'] = get_file_bw_value(fio_write_1M_bw['file_name'])
row += 1
write_bw_excel(fio_write_1M_bw, row)

fio_write_1M_iops = {'name':'100%顺序写模式 iops (次)\n',
        'explain':'顺序写磁盘的每秒写次数\n',
        'file_name':path + 'report/fio_result/1M/write_1M.txt'
        }
fio_write_1M_iops['iops_value'] = get_file_iops_value(fio_write_1M_iops['file_name'])
row += 1
write_iops_excel(fio_write_1M_iops, row)

fio_read_1M_bw = {'name':'100%顺序读模式 bw (KB/s)\n',
        'explain':'顺序读磁盘的吞吐量\n',
        'file_name':path + 'report/fio_result/1M/read_1M.txt'
        }
fio_read_1M_bw['bw_value'] = get_file_bw_value(fio_read_1M_bw['file_name'])
row += 1
write_bw_excel(fio_read_1M_bw, row)

fio_read_1M_iops = {'name':'100%顺序读模式 iops (次)\n',
        'explain':'顺序读磁盘的每秒读次数\n',
        'file_name':path + 'report/fio_result/1M/read_1M.txt'
        }
fio_read_1M_iops['iops_value'] = get_file_iops_value(fio_read_1M_iops['file_name'])
row += 1
write_iops_excel(fio_read_1M_iops, row)

fio_randwrite_1M_bw = {'name':'100%随机写模式 bw (KB/s)\n',
        'explain':'随机写磁盘的吞吐量\n',
        'file_name':path + 'report/fio_result/1M/randwrite_1M.txt'
        }
fio_randwrite_1M_bw['bw_value'] = get_file_bw_value(fio_randwrite_1M_bw['file_name'])
row += 1
write_bw_excel(fio_randwrite_1M_bw, row)

fio_randwrite_1M_iops = {'name':'100%随机写模式 iops (次)\n',
        'explain':'随机写磁盘的每秒写次数\n',
        'file_name':path + 'report/fio_result/1M/randwrite_1M.txt'
        }
fio_randwrite_1M_iops['iops_value'] = get_file_iops_value(fio_randwrite_1M_iops['file_name'])
row += 1
write_iops_excel(fio_randwrite_1M_iops, row)

fio_randread_1M_bw = {'name':'100%随机读模式 bw (KB/s)\n',
        'explain':'随机读磁盘的吞吐量\n',
        'file_name':path + 'report/fio_result/1M/randread_1M.txt'
        }
fio_randread_1M_bw['bw_value'] = get_file_bw_value(fio_randread_1M_bw['file_name'])
row += 1
write_bw_excel(fio_randread_1M_bw, row)

fio_randread_1M_iops = {'name':'100%随机读模式 iops (次)\n',
        'explain':'随机读磁盘的每秒读次数\n',
        'file_name':path + 'report/fio_result/1M/randread_1M.txt'
        }
fio_randread_1M_iops['iops_value'] = get_file_iops_value(fio_randread_1M_iops['file_name'])
row += 1
write_iops_excel(fio_randread_1M_iops, row)

fio_randrw_mixwrite_70_1M_bw = {'name':'写占70%随机混合读写模式 bw (KB/s)\n',
        'explain':'随机混合读写磁盘的吞吐量\n',
        'file_name':path + 'report/fio_result/1M/randrw_mixwrite_70_1M.txt'
        }
fio_randrw_mixwrite_70_1M_bw['bw_value'] = get_file_bw_value(fio_randrw_mixwrite_70_1M_bw['file_name'])
row += 1
write_bw_excel(fio_randrw_mixwrite_70_1M_bw, row)

fio_randrw_mixwrite_70_1M_iops = {'name':'写占70%随机混合读写模式 iops (次)\n',
        'explain':'随机混合读写磁盘的每秒读写次数\n',
        'file_name':path + 'report/fio_result/1M/randrw_mixwrite_70_1M.txt'
        }
fio_randrw_mixwrite_70_1M_iops['iops_value'] = get_file_iops_value(fio_randrw_mixwrite_70_1M_iops['file_name'])
row += 1
write_iops_excel(fio_randrw_mixwrite_70_1M_iops, row)

fio_randrw_mixread_70_1M_bw = {'name':'读占70%随机混合读写模式 bw (KB/s)\n',
        'explain':'随机混合读写磁盘的吞吐量\n',
        'file_name':path + 'report/fio_result/1M/randrw_mixread_70_1M.txt'
        }
fio_randrw_mixread_70_1M_bw['bw_value'] = get_file_bw_value(fio_randrw_mixread_70_1M_bw['file_name'])
row += 1
write_bw_excel(fio_randrw_mixread_70_1M_bw, row)

fio_randrw_mixread_70_1M_iops = {'name':'读占70%随机混合读写模式 iops (次)\n',
        'explain':'随机混合读写磁盘的每秒读写次数\n',
        'file_name':path + 'report/fio_result/1M/randrw_mixread_70_1M.txt'
        }
fio_randrw_mixread_70_1M_iops['iops_value'] = get_file_iops_value(fio_randrw_mixread_70_1M_iops['file_name'])
row += 1
write_iops_excel(fio_randrw_mixread_70_1M_iops, row)
bw.save('fio.xlsx')
